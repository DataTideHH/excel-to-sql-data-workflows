from __future__ import annotations

import csv
import hashlib
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from scripts.data_logic import CUSTOMERS, PRODUCTS, build_orders, revenue_from_order
from scripts.generate_excel_workbook import (
    CANONICAL_EXTERNAL_ATTR,
    FIXED_TIMESTAMP,
    FIXED_W3CDTF_TIMESTAMP,
    OUTPUT_PATH,
    generate_workbook,
)
from scripts.generate_expected_results import EXPECTED_DIR, generate_expected_results
from scripts.generate_sample_data import CSV_LINE_TERMINATOR, RAW_DIR, generate_raw_data

ROOT = Path(__file__).resolve().parents[1]


REQUIRED_CUSTOMER_COLUMNS = {
    "customer_id",
    "customer_name",
    "region",
    "segment",
    "crm_lookup_code",
}
REQUIRED_PRODUCT_COLUMNS = {
    "product_id",
    "product_name",
    "category",
    "list_price",
    "legacy_sku",
}
REQUIRED_ORDER_COLUMNS = {
    "order_id",
    "source_order_ref",
    "order_date",
    "customer_id",
    "product_id",
    "quantity",
    "unit_price",
    "discount_pct",
    "sales_rep",
    "order_status",
    "crm_lookup_code_used",
}


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def test_generated_row_counts() -> None:
    customers = _read_csv(RAW_DIR / "customers.csv")
    products = _read_csv(RAW_DIR / "products.csv")
    orders = _read_csv(RAW_DIR / "orders.csv")

    assert len(customers) == 12
    assert len(products) == 7
    assert len(orders) == 52


def test_required_columns_present() -> None:
    customers = _read_csv(RAW_DIR / "customers.csv")
    products = _read_csv(RAW_DIR / "products.csv")
    orders = _read_csv(RAW_DIR / "orders.csv")

    assert set(customers[0].keys()) == REQUIRED_CUSTOMER_COLUMNS
    assert set(products[0].keys()) == REQUIRED_PRODUCT_COLUMNS
    assert set(orders[0].keys()) == REQUIRED_ORDER_COLUMNS


def test_primary_key_uniqueness_for_required_keys() -> None:
    customers = _read_csv(RAW_DIR / "customers.csv")
    products = _read_csv(RAW_DIR / "products.csv")
    orders = _read_csv(RAW_DIR / "orders.csv")

    assert len({row["customer_id"] for row in customers}) == len(customers)
    assert len({row["product_id"] for row in products}) == len(products)
    assert len({row["order_id"] for row in orders}) == len(orders)


def test_foreign_key_integrity() -> None:
    customers = _read_csv(RAW_DIR / "customers.csv")
    products = _read_csv(RAW_DIR / "products.csv")
    orders = _read_csv(RAW_DIR / "orders.csv")

    customer_ids = {row["customer_id"] for row in customers}
    product_ids = {row["product_id"] for row in products}

    assert all(row["customer_id"] in customer_ids for row in orders)
    assert all(row["product_id"] in product_ids for row in orders)


def test_deterministic_build_orders() -> None:
    orders = build_orders(total_orders=52)

    assert orders[0]["order_id"] == "O001"
    assert orders[0]["source_order_ref"] == "WEB-1001"
    assert orders[17]["source_order_ref"] == "WEB-1017"
    assert orders[35]["source_order_ref"] == "WEB-1017"
    assert orders[-1]["order_id"] == "O052"


def test_revenue_calculation_for_known_row() -> None:
    order = {
        "quantity": "3",
        "unit_price": "100.00",
        "discount_pct": "10",
    }
    assert str(revenue_from_order(order)) == "270.00"


def test_non_unique_lookup_values_exist_for_semantic_demos() -> None:
    customers = _read_csv(RAW_DIR / "customers.csv")
    products = _read_csv(RAW_DIR / "products.csv")

    crm_values = [row["crm_lookup_code"] for row in customers]
    legacy_values = [row["legacy_sku"] for row in products]

    assert crm_values.count("CRM-NORTH-1") > 1
    assert legacy_values.count("LEG-PR-01") > 1


def test_missing_values_exist() -> None:
    orders = _read_csv(RAW_DIR / "orders.csv")
    assert any(row["discount_pct"] == "" for row in orders)
    assert any(row["sales_rep"] == "" for row in orders)


def test_scripts_reference_constants() -> None:
    assert len(CUSTOMERS) == 12
    assert len(PRODUCTS) == 7


def test_generated_csv_files_use_lf_line_endings() -> None:
    generate_raw_data()
    generate_expected_results()
    for path in sorted([*RAW_DIR.glob("*.csv"), *EXPECTED_DIR.glob("*.csv")]):
        data = path.read_bytes()
        assert b"\r\n" not in data, path.as_posix()
        assert data.endswith(CSV_LINE_TERMINATOR.encode("ascii")), path.as_posix()


def test_workbook_generation_is_byte_identical_across_runs(tmp_path: Path) -> None:
    generate_raw_data()
    generate_workbook()
    first_hash = _sha256(OUTPUT_PATH)
    first_copy = tmp_path / "first.xlsx"
    first_copy.write_bytes(OUTPUT_PATH.read_bytes())

    generate_workbook()
    second_hash = _sha256(OUTPUT_PATH)

    assert first_hash == second_hash
    assert first_copy.read_bytes() == OUTPUT_PATH.read_bytes()


def test_workbook_zip_metadata_is_canonical() -> None:
    generate_workbook()

    with zipfile.ZipFile(OUTPUT_PATH, "r") as archive:
        infos = archive.infolist()
        assert [info.filename for info in infos] == sorted(info.filename for info in infos)

        for info in infos:
            assert info.date_time == FIXED_TIMESTAMP
            assert info.compress_type == zipfile.ZIP_STORED
            assert info.create_system == 3
            assert info.external_attr == CANONICAL_EXTERNAL_ATTR
            assert info.flag_bits == 0
            assert info.comment == b""
            assert info.extra == b""

        core_xml = archive.read("docProps/core.xml").decode("utf-8")

    assert FIXED_W3CDTF_TIMESTAMP in core_xml
    assert core_xml.count(FIXED_W3CDTF_TIMESTAMP) == 2


def test_workbook_loads_after_normalization() -> None:
    generate_workbook()
    workbook = load_workbook(OUTPUT_PATH)

    assert workbook.sheetnames == ["Customers", "Products", "Orders", "Exercises", "Summary"]
