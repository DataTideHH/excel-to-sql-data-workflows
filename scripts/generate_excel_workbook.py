"""Build an Excel workbook that demonstrates formula patterns against source CSV data."""

from __future__ import annotations

import csv
import re
import shutil
import zipfile
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
RAW_DIR = ROOT / "data" / "raw"
OUTPUT_PATH = ROOT / "excel" / "excel_sql_workflows.xlsx"
FIXED_TIMESTAMP = (2026, 1, 1, 0, 0, 0)
FIXED_W3CDTF_TIMESTAMP = "2026-01-01T00:00:00Z"
CANONICAL_EXTERNAL_ATTR = 0o100644 << 16
CORE_XML_PATH = "docProps/core.xml"


def _read_csv(path: Path) -> list[dict[str, str]]:
    with path.open("r", newline="", encoding="utf-8") as handle:
        return list(csv.DictReader(handle))


def _write_sheet_with_table(
    workbook: Workbook, title: str, rows: list[dict[str, str]], table_name: str
) -> None:
    sheet = workbook.create_sheet(title=title)
    headers = list(rows[0].keys())
    sheet.append(headers)
    for row in rows:
        sheet.append([row[column] for column in headers])

    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    last_row = sheet.max_row
    last_col = sheet.max_column
    end_col_letter = chr(ord("A") + last_col - 1)
    table = Table(displayName=table_name, ref=f"A1:{end_col_letter}{last_row}")
    style = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)
    sheet.freeze_panes = "A2"


def _build_exercises_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet(title="Exercises")
    sheet.append(["Pattern", "Formula", "Notes"])
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)
    sheet["C1"].font = Font(bold=True)

    rows = [
        (
            "Conditional Logic",
            '=IF([@[order_status]]="Completed","Count","Review")',
            "Equivalent intent: SQL CASE / DAX IF",
        ),
        (
            "Lookup",
            '=XLOOKUP([@[customer_id]],Customers[customer_id],Customers[region],"Missing")',
            "XLOOKUP returns one value; SQL LEFT JOIN may duplicate rows with non-unique keys",
        ),
        (
            "Conditional Sum",
            '=SUMIFS(Orders[quantity],Orders[order_status],"Completed")',
            "Equivalent intent: SUM with WHERE",
        ),
        (
            "Conditional Count",
            '=COUNTIFS(Orders[order_status],"Completed",Orders[sales_rep],"Alex Kim")',
            "Equivalent intent: COUNT with predicates",
        ),
        (
            "Distinct Values",
            "=UNIQUE(Customers[region])",
            "Equivalent intent: DISTINCT / VALUES",
        ),
        (
            "Running Total",
            "=SUM($K$2:K2)",
            "Requires deterministic sort order",
        ),
        (
            "Rank",
            "=RANK.EQ([@[unit_price]],Orders[unit_price],0)",
            "Compare to SQL RANK/DENSE_RANK/ROW_NUMBER",
        ),
        (
            "Missing Value Handling",
            '=IFERROR([@[unit_price]]*[@[quantity]],0)',
            "Contrast blanks, empty strings, NULL, null and BLANK()",
        ),
        (
            "Duplicate Detection",
            '=COUNTIF(Orders[source_order_ref],[@[source_order_ref]])',
            "Values greater than 1 indicate duplicates",
        ),
    ]

    for row in rows:
        sheet.append(list(row))

    sheet.column_dimensions["A"].width = 24
    sheet.column_dimensions["B"].width = 58
    sheet.column_dimensions["C"].width = 76
    sheet.freeze_panes = "A2"


def _build_summary_sheet(workbook: Workbook, orders_count: int) -> None:
    sheet = workbook.create_sheet(title="Summary")
    sheet.append(["Metric", "Formula / Value"])
    sheet["A1"].font = Font(bold=True)
    sheet["B1"].font = Font(bold=True)

    summary_rows = [
        ("Order Rows", str(orders_count)),
        (
            "Total Revenue",
            "=SUMPRODUCT(Orders[quantity],Orders[unit_price],1-(Orders[discount_pct]/100))",
        ),
        ("Completed Orders", '=COUNTIFS(Orders[order_status],"Completed")'),
        (
            "Distinct Customers",
            "=COUNTA(UNIQUE(Orders[customer_id]))",
        ),
        (
            "Top Region (by completed revenue)",
            "Build using PivotTable filter or SORTBY over a region summary range",
        ),
    ]

    for row in summary_rows:
        sheet.append(list(row))

    sheet.column_dimensions["A"].width = 36
    sheet.column_dimensions["B"].width = 72


def _normalize_core_properties(xml_bytes: bytes) -> bytes:
    xml_text = xml_bytes.decode("utf-8")
    xml_text = re.sub(
        r"<dcterms:created xsi:type=\"dcterms:W3CDTF\">.*?</dcterms:created>",
        (
            "<dcterms:created xsi:type=\"dcterms:W3CDTF\">"
            f"{FIXED_W3CDTF_TIMESTAMP}"
            "</dcterms:created>"
        ),
        xml_text,
        count=1,
    )
    xml_text = re.sub(
        r"<dcterms:modified xsi:type=\"dcterms:W3CDTF\">.*?</dcterms:modified>",
        (
            "<dcterms:modified xsi:type=\"dcterms:W3CDTF\">"
            f"{FIXED_W3CDTF_TIMESTAMP}"
            "</dcterms:modified>"
        ),
        xml_text,
        count=1,
    )
    return xml_text.encode("utf-8")


def _canonical_zip_info(name: str, payload: bytes) -> zipfile.ZipInfo:
    info = zipfile.ZipInfo(filename=name, date_time=FIXED_TIMESTAMP)
    info.compress_type = zipfile.ZIP_STORED
    info.create_system = 3
    info.external_attr = CANONICAL_EXTERNAL_ATTR
    info.flag_bits = 0
    info.comment = b""
    info.extra = b""
    info.internal_attr = 0
    info.file_size = len(payload)
    info.CRC = zipfile.crc32(payload) & 0xFFFFFFFF
    return info


def _normalize_xlsx_zip(path: Path) -> None:
    """Rewrite XLSX zip entries with canonical metadata and byte-stable storage."""
    temp_path = path.with_name(f"{path.stem}.canonical{path.suffix}")
    with zipfile.ZipFile(path, "r") as source:
        entries: list[tuple[str, bytes]] = []
        for name in sorted(source.namelist()):
            payload = source.read(name)
            if name == CORE_XML_PATH:
                payload = _normalize_core_properties(payload)
            entries.append((name, payload))

    with zipfile.ZipFile(temp_path, "w", compression=zipfile.ZIP_STORED) as target:
        for name, payload in entries:
            target.writestr(_canonical_zip_info(name, payload), payload)

    shutil.move(temp_path, path)


def generate_workbook() -> None:
    customers = _read_csv(RAW_DIR / "customers.csv")
    products = _read_csv(RAW_DIR / "products.csv")
    orders = _read_csv(RAW_DIR / "orders.csv")

    workbook = Workbook()
    workbook.remove(workbook.active)
    workbook.properties.created = datetime(*FIXED_TIMESTAMP)
    workbook.properties.modified = datetime(*FIXED_TIMESTAMP)
    workbook.properties.creator = "excel-to-sql-data-workflows"
    workbook.properties.lastModifiedBy = "excel-to-sql-data-workflows"

    _write_sheet_with_table(workbook, "Customers", customers, "Customers")
    _write_sheet_with_table(workbook, "Products", products, "Products")
    _write_sheet_with_table(workbook, "Orders", orders, "Orders")
    _build_exercises_sheet(workbook)
    _build_summary_sheet(workbook, orders_count=len(orders))

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(OUTPUT_PATH)
    _normalize_xlsx_zip(OUTPUT_PATH)


if __name__ == "__main__":
    generate_workbook()
    print(f"Generated workbook at {OUTPUT_PATH}")
